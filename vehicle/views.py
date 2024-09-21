from rest_framework import generics 
from rest_framework.response import Response
import json
from .database_opertaions import *
from .utils import *
from .serializers import *
from rest_framework.views import APIView
from datetime import  date
from knox.models import AuthToken
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from django.contrib.auth.models import Group
from rest_framework.authtoken.models import Token
from rest_framework.parsers import MultiPartParser
from .permissions import *
from database.models import *
from rest_framework.throttling import UserRateThrottle , AnonRateThrottle
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views.decorators.cache import cache_page
from django.db.models import Sum
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font 
from django.http import HttpResponse
from django.db.models import Prefetch
from django.utils import timezone



class UserRegister(generics.GenericAPIView):
    serializer_class = RegisterSerializer
    # permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]      

    def post(self, request):
        serializer = RegisterSerializer(data=request.data)
        
        if serializer.is_valid():
            try:
                group_name = serializer.validated_data.get('groups')
                usergroup = Group.objects.get(name= group_name)
            
                user = serializer.save()
        
                if usergroup:
                    user.groups.add(usergroup)
                

                return Response({'message': 'Registration Successfull' , 
                                    "status" : "success" }, status=200)
            except:
                return Response({'message': 'Please select any one group',
                                 "status" : "error"}, status=400)
        else:
            key, value =list(serializer.errors.items())[0]
            try:
                key , value = list(value[0].items())[0]
                error_message =  str(value[0])
            except Exception as e:
                try:
                     error_message = str(key) + " ," +str(value[0])
                except:
                    key , value = list(value[1].items())[0]
                    error_message = str(key) + " ," +str(value[0])
            
            return Response({'message': error_message, 
                             "status" : "error"}, status=400)

class LoginView(generics.GenericAPIView):
    serializer_class = LoginSerializer
    def post(self, request):
        
        serializer = LoginSerializer(data=request.data)

        if serializer.is_valid():
            user_data = serializer.validated_data
            group = user_data.groups.values_list("name", flat=True)[0]
            if serializer is not None:
                _, token = AuthToken.objects.create(serializer.validated_data)
                return Response({
                                'Token':token,
                                'status' : 'success' , 
                                'message' : 'Login successful' ,
                                'username': user_data.username,
                                'group' : group, 
                                    } , status= 200)
        else:
            key, value =list(serializer.errors.items())[0]
            error_message =value[0]
            return Response({'status': 'error',
                            'message': error_message} , status=status.HTTP_400_BAD_REQUEST)

class LogoutView(generics.GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = logoutSerializer

    def post(self, request):
        try:
            token = request.data.get('token')

            token = Token.objects.get(key=token)
            token.delete()
            return Response({"message": "User logged out successfully"}, status=205)
        except Token.DoesNotExist:
            return Response({"message": "Invalid refresh token"}, status=400)


class PostMasterDeviceData(APIView):
    def get(self , request):
        refresh_token = refresh_access_token()
        response = get_device_Data(refresh_token)
    
        if response.status_code == 200:
            try:
                devices_data = json.loads(response.content).get('data')
                # print(devices_data)

                master_data_list = []
                for data in devices_data:
                    # if "deviceDetails" in data:
                    #         create_device_object(data)
                    device_id = data.get("id")
                    device_instances = devices.objects.filter(device_id=device_id).first()
                    
                    if device_instances:
                        master_data_list.append(create_master_device_details(device_instances , data))
                    # else:
                    #     if "deviceDetails" in data:
                    #         create_device_object(data)

                # print(master_data_list)    
                master_object = MasterDeviceDetails.objects.bulk_create(master_data_list)

                return Response({"message": "device data created successfully"}, status=200)
            except Exception as e:
                return Response({"message": str(e)})
        else:
            return Response(response.content, status=response.status_code)
 
   
class ViewDeviceAllDetails(APIView):
    permission_classes = [IsAuthenticated , IsAchargeZone  | IsBattery_IQ | IsAmnex ]
    throttle_classes = [UserRateThrottle , AnonRateThrottle]
    """
    This function is used to filter the queryset based on the 'name' query parameter.
    If 'name' is provided, it filters the devices with names containing the 'name'.
    If 'name' is not provided, it returns all the devices.

    Parameters:
    self (ViewDeviceDetails): The instance of the ViewDeviceDetails class.

    Returns:
    queryset (QuerySet): The filtered queryset of devices.
    """
    def get_queryset(self):
        
        name = self.request.query_params.get('name')
        queryset = devices.objects.none()
        if name:
            queryset |= devices.objects.filter(name__icontains='MBMT')
        else:
           queryset |= devices.objects.all()
            
        return queryset
    
    """
    This function is used to retrieve and serialize device details.
    It fetches all related device details using prefetch_related and then serializes them.
    If no device details are found, it returns an empty response.

    Parameters:
    self (ViewDeviceDetails): The instance of the ViewDeviceDetails class.
    request (Request): The request object containing query parameters.

    Returns:
    Response: A response object containing serialized device details.
    """
    def get(self, request):

        data_list = []
        all_devices = self.get_queryset()
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')

        if all_devices:
            
            for device in all_devices:
                today = date.today()
                if start_date and end_date:
                    try:
                        master_data_list = MasterDeviceDetails.objects.filter(device_id = device , created_at__range=[start_date , end_date]).latest("created_at")
                    except:
                        continue
                else:
                    try:
                        master_data_list = MasterDeviceDetails.objects.filter(device_id = device , created_at__date=today).latest("created_at")
                    except:
                        continue

                data_list_serializer = DataListSerializer(master_data_list).data
                data_list.append({
                    'data' : data_list_serializer
                })

        if len(data_list) ==0:
            return Response({'status': 'error' , 'message': 'No data available'} , status= 200)
        else:
         
             return Response(data_list)
             
class ViewAllMBMTDeviceDetails(generics.GenericAPIView):
    # permission_classes = [IsAuthenticatedOrReadOnly , IsMBMT | IsAmnex ]
    throttle_classes = [AnonRateThrottle]


    """
    This function is used to filter the queryset based on the 'name' query parameter.
    If 'name' is provided, it filters the devices with names containing the 'name'.
    If 'name' is not provided, it returns all the devices.

    Parameters:
    self (ViewDeviceDetails): The instance of the ViewDeviceDetails class.

    Returns:
    queryset (QuerySet): The filtered queryset of devices.
    """
    def get_queryset(self):
        
        name = self.request.query_params.get('name')
        queryset = devices.objects.none()
       
        if name:
            queryset |= devices.objects.filter(name__icontains='MBMT')
        else:
            queryset |= devices.objects.filter(name__icontains='MBMT')
            
        return queryset
    

    """
    This function is used to retrieve and serialize device details.
    It fetches all related device details using prefetch_related and then serializes them.
    If no device details are found, it returns an empty response.

    Parameters:
    self (ViewDeviceDetails): The instance of the ViewDeviceDetails class.
    request (Request): The request object containing query parameters.

    Returns:
    Response: A response object containing serialized device details.
    """
    def get(self, request):
        data_list = []
        all_devices = self.get_queryset()
        if all_devices:
            for device in all_devices:
                device_details_serailizer = deviceDetailsSerialiser(device).data
                today = date.today()
                try:
                    device_location = MasterDeviceDetails.objects.filter(device_id=device, created_at__date=today).latest("created_at")
                    
                    device_location_serializer = MBMTDeviceLocationSerializer(device_location).data
                except:
                    continue

                data_list.append({
                    'device_details' : device_details_serailizer,
                    'device_location': device_location_serializer,
                    
                    })
        else: 
            return Response({'status': 'error' ,
                              'message': 'No data available'} , status= 200)
         
        return Response(data_list)
    
class ViewAmnexDeviceDetails(APIView):
    permission_classes = [IsAuthenticated , IsAmnex | IsBattery_IQ]
    throttle_classes = [UserRateThrottle , AnonRateThrottle]
    def get_queryset(self):

        queryset = devices.objects.filter(name__icontains='MBMT')
            
        return queryset
    """
    This function is used to retrieve and serialize device details.
    It fetches all related device details using prefetch_related and then serializes them.
    If no device details are found, it returns an empty response.

    Parameters:
    self (ViewDeviceDetails): The instance of the ViewDeviceDetails class.
    request (Request): The request object containing query parameters.

    Returns:
    Response: A response object containing serialized device details.
    """
    def get(self, request):
        data_list = []
        all_devices = self.get_queryset()
        

        if all_devices:
          
            for device in all_devices:
                device_details_serailizer = deviceDetailsSerialiser(device).data
                today = date.today()
            
                try:
                    master_data_list = MasterDeviceDetails.objects.filter(device_id = device ,
                                                                          created_at__date=today ).latest("created_at")
                    # select SUM(dm."totalRegenerationEnergy") from public.database_masterdevicedetails dm 
                    # where  dm.device_id = 36 and DATE(dm.created_at) = '2024-08-17';

                    
                    data_list_serializer = DataListSerializer(master_data_list).data

                    # total_Today_RegenerationEnergy = MasterDeviceDetails.objects.filter(device_id=device,
                    #         created_at__date=today).aggregate(total_energy=Sum('totalRegenerationEnergy'))['total_energy']

                except:
                    continue
                data_list.append({
                    'device_details' : device_details_serailizer,
                    'data' : data_list_serializer,
                    #  'total_energy': total_Today_RegenerationEnergy

                })
        else: 
            return Response({'status': 'error' ,
                              'message': 'No data available'} , status= 200)
        return Response(data_list)
    
class GetDeviceParametersDetails(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = deviceDetailsSerialiser
    queryset = devices.objects.all()


    @method_decorator(cache_page(60))
    def dispatch(self, *args, **kwargs):
        return super(GetDeviceParametersDetails, self).dispatch(*args, **kwargs)


# def get_devices_details_view(query_params=None , user_group = None):
#     try:
#         data_list = []
#         if query_params:
#             names = query_params.get('name')
                 
#             for name in names:
#                 devices_list = devices.objects.filter(name__iexact=name)
#                 for device in devices_list:
#                     device_details_serailizer = deviceDetailsSerialiser(device).data
#                     today = date.today()
#                     try:
#                         master_data_list = MasterDeviceDetails.objects.filter(device_id = device , 
#                                                                               created_at__date=today).latest("created_at")
#                         data_list_serializer = DataListSerializer(master_data_list).data
#                     except:
#                         continue
#                     data_list.append({
#                         'device_details' : device_details_serailizer,
#                         'data' : data_list_serializer
#                     })
#         return data_list
#     except Exception as e:
#         return Response({'status': 'error' ,
#                           'message': str(e)} , status= 200)


def get_devices_details_view(query_params=None, user_group=None):
    try:
        if not query_params or 'name' not in query_params:
            return []

        names = query_params.get('name', [])
        today = timezone.now().date()

        devices_queryset = (
            devices.objects.filter(name__in=names)
            .prefetch_related(
                Prefetch(
                    'master_device_id',
                    queryset=MasterDeviceDetails.objects.filter(created_at__date=today),
                    to_attr='today_details'
                )
            )
        )
        data_list = []
        for device in devices_queryset:
            device_details = deviceDetailsSerialiser(device).data
            
            if device.today_details:
                latest_details = max(device.today_details, key=lambda x: x.created_at)
                data_list_serializer = DataListSerializer(latest_details).data
                
                data_list.append({
                    'device_details': device_details,
                    'data': data_list_serializer
                })

        return data_list

    except Exception as e:
        return Response({'status': 'error', 'message': str(e)}, status=500)

class GettimeRangedateData(generics.GenericAPIView):
    def get( self, request , device_name , start_date ,end_date ):
        try:
            data_list = []
            uber_devices = devices.objects.filter(name = device_name)
            for device in uber_devices:
                device_details_serailizer = deviceDetailsSerialiser(device).data
                try:
                    master_data_list = MasterDeviceDetails.objects.filter(device_id = device , created_at__range=(start_date, end_date))
                    data_list_serializer = DataListSerializer(master_data_list , many = True).data
                except:
                    continue
                data_list.append({
                    'device_details' : device_details_serailizer,
                    'data' : data_list_serializer
                })
            return Response (data_list)
        except:
            return Response({'status': 'error' , 'message': 'No data available'} , status= 200)


# Routes and Stop APIs


class GetNoidaExtenToIncedointellectRouteView(generics.GenericAPIView):
    serializer_class = NoidaExtenToIncedointellectRouteSerializer
    def get( self, request):

        route = NoidaExtenToIncedointellectRoute.objects.all()
        data = self.get_serializer(route , many = True).data
        return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': data},
                            status= status.HTTP_200_OK)
    

class GetNoidaExtenToIncedointellectStopsView(generics.GenericAPIView):
    serializer_class = NoidaExtenToIncedointellectStopsSerializer 
    def get( self, request):

        route = NoidaExtenToIncedointellectStops.objects.all()
        data = self.get_serializer(route , many = True).data
        return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': data},
                            status= status.HTTP_200_OK)
    

class GetRouteNo15BusStopsView(generics.GenericAPIView):
    serializer_class = GetRouteNo15BusStopsSerializer 
    def get( self, request):

        route = RouteNo15BusStops.objects.all()
        data = self.get_serializer(route , many = True).data
        return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': data},
                            status= status.HTTP_200_OK)
    

class GetRouteNo15View(generics.GenericAPIView):
    serializer_class = GetRoute15Serializer 
    def get( self, request):

        route = RouteNo15.objects.all()
        data = self.get_serializer(route , many = True).data
        return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': data},
                            status= status.HTTP_200_OK)

class GetChargingStationView(generics.GenericAPIView):
    serializer_class = GetChargingStationSerializer

    def get( self, request):

        route = Chargingsation.objects.all()
        data = self.get_serializer(route , many = True).data
        return Response({'status': 'success',
                            'message' : 'data was successfully fetched',
                            'data': data},
                            status= status.HTTP_200_OK)


class Get_totalRegenerationEnergy(generics.GenericAPIView):
    parser_classes = [MultiPartParser]
    serializer_class = Get_totalRegenerationEnergySerializer

    def get_queryset(self):
        return devices.objects.filter(name__icontains='MBMT')
    
    def get(self, request):
        data_list = []
        all_devices = self.get_queryset()
        serializer = self.get_serializer(data = request.data)
        if serializer.is_valid():
            date = serializer.validated_data.get('date')
        

            if all_devices:
                for device in all_devices:
                    device_details_serializer = deviceDetailsSerialiser(device).data

                    total_Today_RegenerationEnergy = MasterDeviceDetails.objects.filter(
                        device_id=device,
                        created_at__date=date
                    ).aggregate(total_energy=Sum('totalRegenerationEnergy'))['total_energy']
                    
                    data_list.append({
                        'registrationNumber': device_details_serializer.get("registrationNumber"),
                        'name': device_details_serializer.get("name"),
                        'date': date, 
                        'RegenerationEnergy': total_Today_RegenerationEnergy
                    })

            # Create an Excel workbook and add a worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Total Regeneration Energy"

            # Define the header
            headers = ['Vehicle Number', 'Name', 'Date', 'Regeneration Energy']
            ws.append(headers)
            for col_num, header in enumerate(headers, 1):
                col_letter = get_column_letter(col_num)
                ws[col_letter + '1'].font = Font(bold=True)

            # Append data to the worksheet
            for item in data_list:
                ws.append([item['registrationNumber'],item['name'], item['date'], item['RegenerationEnergy']])

            # Prepare the response to return the Excel file
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="total_regeneration_energy_{datetime.datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

            # Save the workbook to the response
            wb.save(response)

            return response
        else:
            key, value =list(serializer.errors.items())[0]
            error_message = key + ", " + value[0]
            return Response({'status': 'error','message': error_message}, status=status.HTTP_400_BAD_REQUEST)
